import win32com.client
import sys
import os, shutil, errno
import xlrd
#import xlsxwriter
import openpyxl
from openpyxl import utils
import re
#import matplotlib.pyplot as plt
from pandas import Series, DataFrame
import pandas as pd

#from PyQt4 import QtCore, QtGui #in the pyqt4 tutorials
#from PyQt5 import QtCore, QtGui, QtWidgets #works for pyqt5
#from PyQt5.QtWidgets import QtGui, QtCore
#from PyQt5.QtWidgets import QtGui, QtCore

import sys

import string

global xls_IMMR_Task_List
global xls_IMMR_ABM_Report
global xls_BJ_Effort

def check_immr_untracked_task():
    print('check_immr_untracked_task\n')
    wk_task_list = xlrd.open_workbook(xls_IMMR_Task_List)
    sheet_task_list = wk_task_list.sheet_by_index(0)

    wk_abm_report = xlrd.open_workbook(xls_IMMR_ABM_Report)
    sheet_abm_task = wk_abm_report.sheet_by_name('Report')

    #get 2nd column from ABM Task
    setTaskABM = set()
    taskABM = []
    for rownum in range(1, sheet_abm_task.nrows):
        taskABM.append(str(sheet_abm_task.cell(rownum, 2).value))
        temp = taskABM[-1]         #some_list[-n] syntax gets the nth-to-last element
        #print("[%s]" % temp)

        #if( temp.startswith('1AR000')):
        #    temp = temp[6:]
        temp = re.sub('1AR0*0','',temp)
        temp = re.sub(' *', '', temp)
        if( temp.endswith('.0')):
            temp = temp[:len(temp)-2]

        taskABM[-1] = temp
        setTaskABM|= {temp}

    #get 1st column from CQ Task List
    taskCQ = []
    for rownum in range(1, sheet_task_list.nrows):
        t = (str(sheet_task_list.cell(rownum,0).value),
             str(sheet_task_list.cell(rownum,1).value),
             str(sheet_task_list.cell(rownum, 2).value),
             str(sheet_task_list.cell(rownum, 7).value)
             )
        taskCQ.append(list(t))
        temp = taskCQ[-1][0]
        temp = re.sub('1AR0*0', '', temp)
        #if( temp.startswith('1AR000')):
        #    temp = temp[6:]

        taskCQ[-1][0] = temp

    print('\ntask CQ' + str(taskCQ))

    task_in_y2016 = set(
        [
            '13387',
            '815',
            '1357',
            '1327',
            '872',
            '1183',
            '739',
            '746',
            '858',
            '13329',
            '13328',
            '6956',
            '7062',
            '7074',
            '7074',
            '7074',
            '7074',
            '7074',
            '7074',
            '7132',
            '7132',
            '7136',
            '7136',
            '7136',
            '7136',
            '7137',
            '7137',
            '7137',
            '7203',
            '7203',
            '7232',
            '7232',
            '7232',
            '7232',
            '7371',
            '7372',
            '7373',
            '7374',
            '7376',
            '7482',
            '7535',
            '7591',
            '7592',
            '7594',
            '7662',
            '7663',
            '11619',
            '11625',
            '13264',
            '13344',
            '13369',
            '11622',
            '13368',
            '13377',
            '10796',
            '11626',
            '13346',
            '13179',
            '13189',
            '13322',
            '13337',
            '13361',
            '11590',
            '12693',
            '13363',
            '13323',
            '13333',
            '13341',
            '12565',
            '11621',
            '13385',
            '13499',
            '13598',
            '14315',
            '14045',
            '14047',
            '13597',
            '13413',
            '13427',
            '14323',
            '13484',
            '13502',
            '13560',
            '13864',
            '14005',
            '13423',
            '13487',
            '13535',
            '13555',
        ])

    print('\nsetTaskABM' + str(setTaskABM))
    #print('\ntask_in_y2016' + str(task_in_y2016))

    task_missed = [[]]
    for index in range(len(taskCQ)):
        #rt = str_in_list((taskCQ[index])[0], taskABM)
        #rt = rt or str_in_list((taskCQ[index])[0], task_in_y2016)

        if( (not (taskCQ[index])[0] in setTaskABM) and ( not (taskCQ[index])[0] in task_in_y2016) ):
            task_missed.append(taskCQ[index])
            print("\nmissed task ", taskCQ[index])


def str_in_list(astr, lst):
    rt = False
    for index in range( len(lst)):
        if( lst[index].find(astr) != -1):
            rt = True
            #print("task %s, index %d" % (astr, index))
            break
    #if( not rt):
        #print("missed task %s" % astr)

    return rt
def sync_ABMs():
    print('sync_ABMs\n')
    os.system('Y:\P_CommNav\Projects\IMMR_Airbus\Snapshot_Quantum\Sync_ABM.bat')
    #src = 'Q:\IMMR Airbus\ProjectHandbook\ABMs\ABMS\HSW'
    #dst = 'Y:\\P_CommNav\\Projects\\IMMR_Airbus\\Snapshot_Quantum\\IMMR Airbus\\ProjectHandbook\\ABMs\ABMS\\HSW'
    #try:
    #    shutil.copytree(src, dst)
    #    #shutil.copy(src, dst)
    #except OSError as exc:  # python >2.5
    #    # File already exist
    #    if exc.errno == errno.EEXIST:
    #        shutil.copy(src, dst)
    #    # The dirtory does not exist
    #    if exc.errno == errno.ENOENT:
    #        shutil.copy(src, dst)
    #    else:
    #        raise

def update_ABM_task():
    print('update_ABM_task\n')
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Workbooks.Open(Filename = xls_IMMR_ABM_Report, ReadOnly = 1)
    xl.Application.Run("cmd_Update")
    xl.Workbooks(1).Close(SaveChanges = 0)
    xl.Application.Quit()
    xl = 0

def update_used_hours():
    print('update_used_hours\n')

    xlsSAP = xlrd.open_workbook(xls_BJ_Effort)
    sheetSAP = xlsSAP.sheet_by_index(0)
    #xlsABMReport = openpyxl.Workbook(xls_IMMR_ABM_Report)
    #sheetEffort = xlsABMReport.get_sheet_by_name('Effort')

    lst_WBS = [['AE-00003689-001-0010','IMMR BITE Coding and Test', '0','C2'],
               ['AE-00003689-001-0011','IMMR SW Host Engr Support','0','C3'],
               ['AE-00003689-001-0012','IMMR IO Host SW Req','0','C4'],
               ['AE-00003689-001-0013','IMMR IO HOST Code and Test','0','C5'],
               ['AE-00003689-001-0014','L2 Interactive Req','0','C6'],
               ['AE-00003689-001-0015','L2 Interactive Code','0','C7'],
               ['AE-00003689-001-0016', 'L2 FLS Req', '0','C8'],
               ['AE-00003689-001-0017', 'L2 FLS Code', '0','C9'],
               ['AE-00003689-001-0018', 'IMMR L1A SW SCR', '0','C10'],
               ['AE-00003689-001-0019', 'SW Doc and Analysis', '0','C11'],
               ['AE-00003689-001-0020', 'SW MCDU Bite', '0','C12'],
               ['AE-00003741-001-0005', 'AMQP Dev', '0','C14'],
               ['AE-00004450-002-0001', 'MkII BGA SW', '0','C15'],
               ['AE-00004450-002-0002', 'MkII BGA Sys', '0','C16'],
               ['AE-00003789-002-0002', 'VPD-EPIC Engineering Kits', '0','C17'],
               ['AE-00003924-005-0001', 'AOIP QE CNS',  '0','C18'],
               ['AE-00003924-005-0002', 'AOIP QE SVV',  '0','C19'],
               ['AE-00004056-001-0003', 'AOIP QE CNS-2',    '0','C20'],
               ['AE-00003810-007-0004', 'Finance OEF Tool', '0','C21'],
               ['AE-00004927-001-0002', 'B777 BPV17B',      '0', 'C22'],
               ]
    lst_wbs_effort = []
    #for rownum in range(2, sheetEffort.nrows):
    #    lst_WBS.append(sheetEffort.cell(rownum,2).value)

    lst_sap_effort=[] #SAP export effort list
    for rownum in range(2, sheetSAP.nrows):
        if(len(sheetSAP.cell(rownum,5).value) > 0):
            t = (sheetSAP.cell(rownum,5).value, sheetSAP.cell(rownum,15).value)
            lst_sap_effort.append(list(t))
            for wbs in range(0,len(lst_WBS)):
                if(lst_WBS[wbs][0]== sheetSAP.cell(rownum,5).value ):
                    lst_WBS[wbs][2] = float(lst_WBS[wbs][2]) +  float(sheetSAP.cell(rownum,15).value)
                    break

    #print(lst_WBS)
    for ele in lst_WBS:
        print(ele)

    #write data back to xls
    wt_wkxls = openpyxl.load_workbook(xls_IMMR_ABM_Report, read_only=False, keep_vba=True)
    #wt_sheet = wt_wkxls.get_sheet_by_name(r'Effort')
    wt_sheet = wt_wkxls.get_sheet_by_name("Effort")
    for wbs in range(0,len(lst_WBS)):
        #wt_sheet.cell(lst_WBS[wbs][3]).value = lst_WBS[wbs][2]
        r, c = utils.coordinate_to_tuple(lst_WBS[wbs][3])
        wt_sheet.cell(row=r, column=c).value = lst_WBS[wbs][2]
    wt_wkxls.save(xls_IMMR_ABM_Report)
    wt_wkxls.close()
#def msg_box(str1, str2):
    #QtWidgets.QMessageBox.about("My message box", "Text1 = %s, Text2 = %s" % ('T1', 'T2') )

if __name__ == '__main__':
    try:
        xls_IMMR_Task_List = r'Y:\P_CommNav\Projects\IMMR_Airbus\Snapshot_Quantum\IMMR Airbus\ProjectHandbook\Financials\WeeklyActuals\Wkly_TT_Report\QueryResult.xls'
        xls_IMMR_ABM_Report = r'C:\Users\e427632\Google Drive\Lns\IMMR\ABM_Report.xlsm'
        xls_BJ_Effort = r'Y:\Process\SAP\Effort_2017\CNS_BJ_Effort_By_Now.xlsX'
        xls_TT_Task_Report = r'Y:\P_CommNav\Projects\IMMR_Airbus\Snapshot_Quantum\IMMR Airbus\ProjectHandbook\Financials\WeeklyActuals\Wkly_TT_Report\IMMR_SpendReports.xlsx'

        #os.system(r'cmd.exe copy /Y C:\Users\e427632\Google Drive\Lns\IMMR\ABM_Report.xlsm Q:\IMMR Airbus\ProjectHandbook\ABMs\ABMS\HSW\ABM_Report.xlsm')
        #os.system(r'cmd.exe copy /Y C:\Users\e427632\Google Drive\Lns\IMMR\ABM_Report.xlsm Y:\P_CommNav\Projects\IMMR_Airbus\Snapshot_Quantum\IMMR Airbus\ProjectHandbook\ABMs\ABMS\HSW\ABM_Report.xlsm')

        #sync_ABMs()

        #update_ABM_task()

        check_immr_untracked_task()
        #msg_box('1','2')

        update_used_hours()
    except:
        print("error happens, ", sys.exc_info()[0])

    input('press any key to continue!')